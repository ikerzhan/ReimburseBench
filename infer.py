import os
import sys
import json
import glob
import time
import random
import base64
import mimetypes
import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from docx import Document
from docx.oxml.ns import qn
import fitz  # PyMuPDF
from PIL import Image

# ------------------------------------------------------------
# Constants
# ------------------------------------------------------------
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.bmp', '.webp', '.gif', '.tiff'}

# Image upload methods
IMAGE_UPLOAD_DIRECT = "direct"
IMAGE_UPLOAD_PIL = "pil"

# ------------------------------------------------------------
# Image processing utilities
# ------------------------------------------------------------
def encode_image_direct(image_path: str) -> str:
    """Read image bytes directly and base64 encode, preserving original format."""
    with open(image_path, "rb") as f:
        img_data = base64.b64encode(f.read()).decode("utf-8")
    mime, _ = mimetypes.guess_type(image_path)
    if not mime:
        mime = "image/jpeg"
    return f"data:{mime};base64,{img_data}"

def encode_image_via_pil(image_path: str) -> str:
    """Open image with PIL and convert to PNG before base64 encoding."""
    with Image.open(image_path) as pil_image:
        if pil_image.mode not in ("RGB", "L"):
            pil_image = pil_image.convert("RGB")
        buffer = BytesIO()
        pil_image.save(buffer, format="PNG")
        img_data = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{img_data}"

def encode_image(image_path: str, method: str = IMAGE_UPLOAD_DIRECT) -> str:
    """Encode image using the specified method."""
    if method == IMAGE_UPLOAD_PIL:
        return encode_image_via_pil(image_path)
    return encode_image_direct(image_path)

# ------------------------------------------------------------
# Document to text utilities
# ------------------------------------------------------------
def read_xlsx_to_markdown(file_path: str) -> str:
    wb = load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"## Sheet: {sheet_name}\n")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        col_count = len(header)
        parts.append("| " + " | ".join([str(col) if col is not None else "" for col in header]) + " |")
        parts.append("| " + " | ".join(["---"] * col_count) + " |")
        for row in rows[1:]:
            row_vals = [str(cell) if cell is not None else "" for cell in row]
            if len(row_vals) < col_count:
                row_vals += [""] * (col_count - len(row_vals))
            parts.append("| " + " | ".join(row_vals) + " |")
        parts.append("")
    return "\n".join(parts)

def read_docx_text(file_path: str) -> str:
    """Extract all text from a Word document, including tables (converted to Markdown)."""
    doc = Document(file_path)
    full_text = []
    for element in doc.element.body:
        if element.tag == qn('w:p'):
            from docx.text.paragraph import Paragraph
            para = Paragraph(element, doc)
            text = para.text.strip()
            if text:
                full_text.append(text)
        elif element.tag == qn('w:tbl'):
            from docx.table import Table
            table = Table(element, doc)
            rows = table.rows
            if not rows:
                continue
            header_cells = [cell.text.replace('\n', ' ') for cell in rows[0].cells]
            col_count = len(header_cells)
            full_text.append('| ' + ' | '.join(header_cells) + ' |')
            full_text.append('| ' + ' | '.join(['---'] * col_count) + ' |')
            for row in rows[1:]:
                cells = [cell.text.replace('\n', ' ') for cell in row.cells]
                if len(cells) < col_count:
                    cells += [''] * (col_count - len(cells))
                full_text.append('| ' + ' | '.join(cells[:col_count]) + ' |')
            full_text.append('')
    return '\n'.join(full_text)

def pdf_to_text_blocks(pdf_path: str) -> List[Dict]:
    """Extract text from each page of a PDF. Warns if a page has no extractable text."""
    doc = fitz.open(pdf_path)
    blocks = []
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text()
        if text.strip():
            blocks.append({
                "type": "text",
                "text": f"[PDF Page {page_num+1}]\n{text}"
            })
        else:
            blocks.append({
                "type": "text",
                "text": (
                    f"[PDF Page {page_num+1}: No extractable text. "
                    f"This page may be a scanned image or image-only PDF; "
                    f"text-only models will be unable to read its content.]"
                )
            })
    doc.close()
    return blocks

# ------------------------------------------------------------
# Core: process source files into ordered content blocks
# ------------------------------------------------------------
def process_source_files(work_dir: str, source_files: dict,
                         image_upload_method: str = IMAGE_UPLOAD_DIRECT) -> List[Dict[str, Any]]:
    """
    Returns an ordered list of content blocks in OpenAI API content format:
      - {"type": "text", "text": "..."}
      - {"type": "image_url", "image_url": {"url": "...", "detail": "auto"}}
    Text and image blocks are ordered sequentially to ensure the model
    can correctly associate images with their context labels.
    """
    blocks: List[Dict[str, Any]] = []

    # ---------- Background_information ----------
    bg_files = source_files.get("Background_information", [])
    if bg_files:
        blocks.append({"type": "text", "text": "\n## Background Documents\n"})

    for fname in bg_files:
        fpath = os.path.join(work_dir, fname)
        if not os.path.exists(fpath):
            blocks.append({"type": "text", "text": f"[WARNING] File not found: {fname}\n"})
            continue

        ext = Path(fname).suffix.lower()

        if ext in IMAGE_EXTS:
            blocks.append({
                "type": "text",
                "text": f"\n### Background image: {fname}\n(The image is shown directly below.)"
            })
            blocks.append({
                "type": "image_url",
                "image_url": {"url": encode_image(fpath, image_upload_method), "detail": "auto"}
            })
            blocks.append({"type": "text", "text": f"(End of background image: {fname})\n"})

        elif ext == '.pdf':
            try:
                pdf_blocks = pdf_to_text_blocks(fpath)
                pdf_text = "\n".join(b["text"] for b in pdf_blocks)
                blocks.append({
                    "type": "text",
                    "text": f"\n### File: {fname} (PDF)\n{pdf_text}\n(End of {fname})\n"
                })
            except Exception as e:
                blocks.append({"type": "text", "text": f"[ERROR reading PDF {fname}: {e}]\n"})

        elif ext == '.xlsx':
            try:
                md_table = read_xlsx_to_markdown(fpath)
                blocks.append({
                    "type": "text",
                    "text": f"\n### File: {fname} (Excel)\n{md_table}\n(End of {fname})\n"
                })
            except Exception as e:
                blocks.append({"type": "text", "text": f"[ERROR reading Excel {fname}: {e}]\n"})

        elif ext in ['.doc', '.docx']:
            try:
                doc_text = read_docx_text(fpath)
                blocks.append({
                    "type": "text",
                    "text": f"\n### File: {fname} (Word)\n{doc_text}\n(End of {fname})\n"
                })
            except Exception as e:
                blocks.append({"type": "text", "text": f"[ERROR reading Word {fname}: {e}]\n"})

        else:
            try:
                with open(fpath, 'r', encoding='utf-8') as f:
                    content = f.read()
                blocks.append({
                    "type": "text",
                    "text": f"\n### File: {fname}\n{content}\n(End of {fname})\n"
                })
            except Exception:
                blocks.append({"type": "text", "text": f"[Cannot display {fname}]\n"})

    # ---------- Receipts ----------
    receipts = source_files.get("Receipts", {})
    if receipts and any(k for k in receipts if k != "standard"):
        blocks.append({"type": "text", "text": "\n## Submitted Receipts (grouped by owner)\n"})

    for key, value in receipts.items():
        if key == "standard":
            continue
        attendee = key

        if isinstance(value, str):
            file_list = [value]
        elif isinstance(value, list):
            file_list = value
        else:
            blocks.append({
                "type": "text",
                "text": f"[WARNING] Receipts value for {attendee} is invalid type, skipping.\n"
            })
            continue

        for fname in file_list:
            fpath = os.path.join(work_dir, fname)
            if not os.path.exists(fpath):
                blocks.append({
                    "type": "text",
                    "text": f"[WARNING] Receipt for {attendee} not found: {fname}\n"
                })
                continue

            ext = Path(fname).suffix.lower()

            if ext in IMAGE_EXTS:
                blocks.append({
                    "type": "text",
                    "text": (
                        f"\n### Receipt belonging to **{attendee}** (file: {fname})\n"
                        f"The receipt image for **{attendee}** is shown directly below:"
                    )
                })
                blocks.append({
                    "type": "image_url",
                    "image_url": {"url": encode_image(fpath, image_upload_method), "detail": "auto"}
                })
                blocks.append({
                    "type": "text",
                    "text": f"(End of receipt for **{attendee}**)\n"
                })

            elif ext == '.pdf':
                try:
                    pdf_blocks = pdf_to_text_blocks(fpath)
                    pdf_text = "\n".join(b["text"] for b in pdf_blocks)
                    blocks.append({
                        "type": "text",
                        "text": (
                            f"\n### Receipt belonging to **{attendee}** (PDF file: {fname})\n"
                            f"{pdf_text}\n"
                            f"(End of receipt for **{attendee}**)\n"
                        )
                    })
                except Exception as e:
                    blocks.append({
                        "type": "text",
                        "text": f"[ERROR reading PDF receipt for {attendee}: {e}]\n"
                    })
            else:
                blocks.append({
                    "type": "text",
                    "text": f"[INFO] Receipt {fname} for {attendee} is not an image/PDF, skipped.\n"
                })

    return blocks

# ============================================================
# Prompt construction
# ============================================================

# ---------- Basic system message ----------
SYSTEM_MSG_NO_CONTEXT = (
    "You are a meticulous reimbursement audit assistant. "
    "Now you will receive a reimbursement application for a business expense, which includes the background information of the transaction, "
    "the claimed reimbursement amount, relevant supporting documents and receipts, as well as the company's internal control policies or guidelines. "
    "Your task is to analyze the provided documents, receipts, and policy files, "
    "then determine whether the claimed expense can be reimbursed, calculate the maximum legally reimbursable amount, "
    "and explain the entire reasoning chain step by step."
)

# ---------- System message with expert knowledge (four-stage audit workflow) ----------
SYSTEM_MSG_WITH_CONTEXT = (
    "You are a meticulous reimbursement audit assistant for corporate financial internal control. "
    "You will receive a reimbursement application together with supporting receipts, background documents, "
    "and the company's internal control policy. Your task is to determine whether the claimed expense "
    "can be reimbursed in full, calculate the maximum legally reimbursable amount, "
    "and explain your reasoning step by step.\n\n"

    "## Core Audit Principles\n"
    "As a corporate auditor, you operate under three fundamental principles:\n"
    "- **Substance over form**: judge the economic reality of a transaction, not its surface appearance.\n"
    "- **Verifiable evidence**: every reimbursable amount must be traceable to a specific, valid piece of evidence in the submitted materials. "
    "Do not assume facts that are not in evidence.\n"
    "- **Policy adherence**: apply the company's internal control rules exactly as written. Do not invent rules, "
    "and do not weaken rules to favor the applicant.\n\n"

    "## Standard Audit Workflow\n"
    "A complete reimbursement audit proceeds in four sequential stages. For each receipt, you must verify all "
    "four stages; failure at any stage affects either the receipt's validity or its reimbursable amount.\n\n"

    "**Stage 1 — Authenticity Check (is the receipt itself genuine?)**\n"
    "Verify that each receipt is a legitimate document, not fabricated or tampered with. "
    "Do not dismiss a receipt as fake solely due to minor font or layout inconsistencies; "
    "instead, focus on substantive logical discrepancies, such as "
    "contradictory content, impossible dates, or values that cannot be physically read from the image. "
    "An illegible or unreadable receipt cannot serve as audit evidence. "
    "*Impact*: A receipt that fails authenticity contributes **0** to the reimbursable amount (Step 1 of MAX_LEGAL calculation).\n\n"

    "**Stage 2 — Compliance Check (does the receipt match the claimed expense and policy scope?)**\n"
    "For each receipt, verify that:\n"
    "  (a) it contains all required fields per company policy;\n"
    "  (b) its expense category matches what is being claimed (e.g., a meal receipt should not be claimed under transportation);\n"
    "  (c) its date falls within the valid period of the business activity;\n"
    "  (d) its location is consistent with the business activity's location;\n"
    "  (e) the personnel involved (e.g., attendees of a business meal) match the names and headcount declared in the application;\n"
    "  (f) any other policy-specific requirements (e.g., minimum information for cross-border travel, the validity period of the company policy) are met.\n"
    "*Impact*: A receipt that fails compliance contributes **0** to the reimbursable amount (Step 1 of MAX_LEGAL calculation).\n\n"

    "**Stage 3 — Supporting Documentation Check (are the required ancillary documents present?)**\n"
    "Many expense types require supporting documents beyond the receipts themselves. The specific requirements "
    "are defined in the company's policy. "
    "*Impact*: Missing supporting documents indicate a procedural defect in the application. "
    "This should be raised as a problem. Whether it directly invalidates the underlying receipt's monetary value "
    "depends on whether the company's policy explicitly states such an invalidation; "
    "if the policy does not specify, treat it as a problem to flag but do not automatically reduce MAX_LEGAL.\n\n"

    "**Stage 4 — Amount Integrity Check (is the requested amount mathematically and policy-wise sound?)**\n"
    "For each surviving receipt and for the application as a whole, verify that:\n"
    "  (a) per-receipt or per-day caps are respected (e.g., hotel allowance per night per person);\n"
    "  (b) no receipt is being claimed more than once across this or related applications (no duplicate reimbursement);\n"
    "  (c) the expense is for a legitimate business purpose, not personal use or padding to reach a quota.\n"
    "*Impact*: An over-cap portion is reduced to the cap (Step 2 of MAX_LEGAL calculation). A duplicate or non-business item "
    "should be removed (Step 1).\n\n"
)

# ---------- Shared user intro text ----------
def build_user_intro_text(context: str, required_amount: float, currency: str) -> str:
    return (
        f"Reimbursement Description: {context}\n"
        f"Claimed reimbursement amount: {required_amount} {currency}\n\n"
        "## IMPORTANT INSTRUCTIONS (read carefully BEFORE examining the materials)\n"
        "1. **Receipt ownership**: The person responsible for each receipt is solely determined by the label provided "
        "in the file description (e.g., 'Receipt belonging to **Zhang_Ming**'). "
        "Disregard any names or identity information printed on the receipt image or PDF itself when assigning the expense to an individual. "
        "Always use the person specified in the label.\n"
        "2. **Strict policy adherence**: Apply rules strictly based on the provided policy documents. Do not invent rules. "
        "Do not assume facts that are not in evidence.\n"
        "3. **Attention to detail**: Pay close attention to dates, locations, personnel levels, amount limits, and any "
        "conditional clauses in the policy.\n"
        "4. **Ignore submission date**: Assume all reimbursement applications are submitted within the legal time frame.\n\n"
        "## MATERIALS\n"
        "Below are all the background materials and receipts needed for the audit, provided in order:\n"
    )

# ---------- Shared output format text ----------
SHARED_OUTPUT_FORMAT_TEXT = (
    "\n\n## OUTPUT FORMAT\n"
    "Produce your response in EXACTLY the following format.\n\n"
    "### Definitions\n"
    "Let **CLAIMED** = the reimbursement amount the applicant submitted (given above).\n"
    "Let **MAX_LEGAL** = the maximum amount that would be reimbursable under the policy, computed as follows:\n"
    "  Step 1: Discard any individual receipt that you determine to be entirely invalid based on the "
    "provided materials and policy. These receipts contribute 0.\n"
    "  Step 2: For each remaining valid receipt, cap its reimbursable portion at the applicable policy limit "
    "(e.g., if the per-night hotel cap is 100 and the receipt is 120, only 100 is reimbursable from that receipt).\n"
    "  Step 3: Sum the capped amounts from Step 2. The result is MAX_LEGAL.\n"
    "  Note: MAX_LEGAL is 0 only if every receipt is fully invalid or there is no valid claim portion at all.\n\n"
    "### The FLAG\n"
    "FLAG is **strictly** the comparison result between MAX_LEGAL and CLAIMED:\n"
    "  - FLAG = 1 if MAX_LEGAL >= CLAIMED (the requested amount is fully supported by the materials and policy).\n"
    "  - FLAG = 0 if MAX_LEGAL <  CLAIMED (the requested amount exceeds what the materials and policy allow).\n"
    "FLAG is a deterministic comparison: compute MAX_LEGAL first, then output FLAG accordingly.\n\n"
    "### Required output structure\n"
    "Line 1: `0` or `1` \u2014 the FLAG value defined above.\n"
    "Line 2: A single numeric value \u2014 MAX_LEGAL in the SAME currency as the claim.\n"
    "Line 3: The exact text `--- Reasoning ---`\n"
    "Following lines: A detailed, step-by-step reasoning chain that explains:\n"
    "   (a) Step 1 results: which receipts (if any) are entirely invalid, and why (cite policy clauses).\n"
    "   (b) Step 2 results: for each valid receipt, show the raw amount, the applicable cap, and the capped amount.\n"
    "   (c) Step 3 results: the arithmetic that produces MAX_LEGAL.\n"
    "   (d) The FLAG: explicitly compare MAX_LEGAL vs CLAIMED and state the resulting FLAG.\n"
    "   (e) Any discrepancies, missing documents, policy violations, or calculation concerns you identified.\n\n"
    "After the reasoning, add a line with the exact text `--- Problems (if any) ---`\n\n"
    "Then list each problem identified during the reasoning stage that concerns the submitted reimbursement, "
    "one per line, in the format: `Problem: <description>`. "
    "Each problem description should clearly state (i) what the problem is, (ii) which specific receipt, document, "
    "or person it concerns, and (iii) the numeric or factual detail that demonstrates it "
    "(e.g., the amount, the date, or the policy clause being violated).\n"
    "If no problems exist, write only: `No problem`.\n\n"
    "Important: The reasoning must include thorough citations of policy clauses and explicit arithmetic. Do not skip any step."
)

# ---------- Unified build_messages entry point ----------
def build_messages(context: str,
                   content_blocks: List[Dict[str, Any]],
                   required_amount: float, currency: str,
                   expert_prompt: bool) -> List[Dict]:
    """
    Build messages for model inference.
    Expert knowledge is injected via system message selection.
    """
    system_content = SYSTEM_MSG_WITH_CONTEXT if expert_prompt else SYSTEM_MSG_NO_CONTEXT
    system_msg = {"role": "system", "content": system_content}

    intro_text = build_user_intro_text(context, required_amount, currency)
    output_format_text = SHARED_OUTPUT_FORMAT_TEXT

    user_content: List[Dict[str, Any]] = [{"type": "text", "text": intro_text}]
    user_content.extend(content_blocks)
    user_content.append({"type": "text", "text": output_format_text})

    return [system_msg, {"role": "user", "content": user_content}]

# ------------------------------------------------------------
# Placeholder for model inference
# ------------------------------------------------------------
def call_model_with_retry(messages: List[Dict], **kwargs) -> Any:
    """
    PLACEHOLDER: Replace this function with your own model inference implementation.

    Expected input:
      - messages: list of dicts in OpenAI chat format
      - kwargs: additional parameters (e.g., max_tokens, temperature)

    Expected output:
      - An object with fields:
        .content: str (the generated response text)
        .reasoning_content: Optional[str] (chain-of-thought, if available)
        .usage: Optional[Dict] with keys like input_tokens, output_tokens, total_tokens

    This function should handle retries and error handling internally.
    """
    raise NotImplementedError(
        "Model inference is not implemented. "
        "Please implement call_model_with_retry() to integrate with your model API."
    )

# ------------------------------------------------------------
# Incremental saving (JSONL)
# ------------------------------------------------------------
def append_jsonl(filepath: str, entry: Dict[str, Any]) -> None:
    with open(filepath, 'a', encoding='utf-8') as f:
        f.write(json.dumps(entry, ensure_ascii=False) + '\n')

def jsonl_to_json(jsonl_path: str, json_path: str) -> None:
    entries = []
    if not os.path.exists(jsonl_path):
        return
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except Exception as e:
                print(f"[WARNING] Skipping unparseable JSONL line: {e}")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(entries, f, indent=2, ensure_ascii=False)

# ------------------------------------------------------------
# Main entry point
# ------------------------------------------------------------
def run_tests(root_dir: str, model_list: List[str], api_key: str, output_file: str,
              max_tokens: int = 8096, enable_thinking: bool = True,
              expert_prompt: bool = False,
              image_upload_method: str = IMAGE_UPLOAD_DIRECT):
    """
    expert_prompt=False  -> v1_1_no_context (basic system prompt)
    expert_prompt=True   -> v1_1_with_context (four-stage audit workflow)

    image_upload_method:
        IMAGE_UPLOAD_DIRECT -> read bytes directly (preserves original format)
        IMAGE_UPLOAD_PIL    -> open with PIL and convert to PNG
    """
    prompt_version = "v1_1_with_context" if expert_prompt else "v1_1_no_context"

    workflows = []
    for sub in sorted(os.listdir(root_dir)):
        sub_path = os.path.join(root_dir, sub)
        if not os.path.isdir(sub_path):
            continue
        task_files = glob.glob(os.path.join(sub_path, "*task_description*"))
        if not task_files:
            print(f"Skipping {sub}: no task_description file")
            continue
        task_file = task_files[0]
        with open(task_file, 'r', encoding='utf-8') as f:
            task_data = json.load(f)
        workflows.append({
            "id": task_data.get("id", sub),
            "path": sub_path,
            "task_data": task_data
        })

    jsonl_path = output_file.replace('.json', '.jsonl')
    if os.path.exists(jsonl_path):
        print(f"[INFO] Existing JSONL file found: {jsonl_path}, new results will be appended.")

    for wf in workflows:
        wf_id = wf["id"]
        wf_path = wf["path"]
        task_data = wf["task_data"]
        context = task_data.get("context", "")
        required_amount = task_data.get("required_amount", 0)
        currency = task_data.get("currency", "CNY")
        source_files = task_data.get("source_files", {})

        content_blocks = process_source_files(wf_path, source_files,
                                              image_upload_method=image_upload_method)
        messages = build_messages(context, content_blocks,
                                  required_amount, currency,
                                  expert_prompt=expert_prompt)

        for model in model_list:
            print(f"Testing {wf_id} -> model {model} [prompt={prompt_version}, img={image_upload_method}] ...")
            response_text = ""
            reasoning_text = ""
            error_msg: Optional[str] = None
            usage_info: Optional[Dict[str, Any]] = None
            t_start = time.time()

            # --------------------------------------------------------
            # PLACEHOLDER: Replace this block with your model inference logic
            # --------------------------------------------------------
            try:
                # Example:
                # result = call_model_with_retry(messages, max_tokens=max_tokens)
                # response_text = result.content
                # reasoning_text = getattr(result, 'reasoning_content', '') or ''
                # usage_info = result.usage

                raise NotImplementedError(
                    "Model inference is not implemented. "
                    "Replace the placeholder block in run_tests() "
                    "with your own model calling logic."
                )
            except NotImplementedError:
                raise
            except Exception as e:
                error_msg = str(e)
            # --------------------------------------------------------

            elapsed = time.time() - t_start

            result_entry: Dict[str, Any] = {
                "workflow_id": wf_id,
                "model": model,
                "prompt_version": prompt_version,
                "enable_thinking": enable_thinking,
                "image_upload_method": image_upload_method,
                "generated_answer": response_text,
                "reasoning_content": reasoning_text if reasoning_text else None,
                "error": error_msg,
                "usage": usage_info,
                "latency_seconds": round(elapsed, 3),
                "timestamp": datetime.datetime.now().isoformat(),
            }

            append_jsonl(jsonl_path, result_entry)

            if error_msg:
                print(f"  -> Error: {error_msg} (elapsed {elapsed:.1f}s)")
            else:
                input_tok = usage_info.get("input_tokens") if usage_info else "N/A"
                output_tok = usage_info.get("output_tokens") if usage_info else "N/A"
                print(f"  -> Done | answer length: {len(response_text)} | "
                      f"reasoning length: {len(reasoning_text)} | "
                      f"input tokens: {input_tok} | output tokens: {output_tok} | "
                      f"elapsed: {elapsed:.1f}s")

    jsonl_to_json(jsonl_path, output_file)
    print(f"\nAll results saved incrementally to: {jsonl_path}")
    print(f"Merged to JSON array: {output_file}")

if __name__ == "__main__":
    load_dotenv()

    root_dir = r"ReimburseBench_v1\workflow"
    models = ["your-model-name"]
    output_path = f"output/{datetime.datetime.now().strftime('%y%m%d%H%M%S')}_results.json"
    MAX_TOKENS = 98304
    ENABLE_THINKING = True
    USE_EXPERT_PROMPT = True
    IMAGE_UPLOAD_METHOD = IMAGE_UPLOAD_DIRECT

    api_key = os.environ.get("YOUR_API_KEY")
    if not api_key:
        print("Please set the environment variable YOUR_API_KEY")
        sys.exit(1)
    run_tests(root_dir, models, api_key, output_path,
              max_tokens=MAX_TOKENS, enable_thinking=ENABLE_THINKING,
              expert_prompt=USE_EXPERT_PROMPT,
              image_upload_method=IMAGE_UPLOAD_METHOD)
